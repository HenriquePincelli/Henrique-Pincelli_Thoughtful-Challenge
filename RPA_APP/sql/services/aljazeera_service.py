from RPA_APP.sql.services.service import Service
from RPA_APP.sql.models.aljazeera_model import AljazeeraModel
from RPA_APP.extensions import db
from datetime import datetime
from sqlalchemy import desc, and_
from sqlalchemy.exc import IntegrityError
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import json
import sys
import pytz
import re
import os
import openpyxl


# >>>>>>>>>Money patterns for regex>>>>>>>>>
money_patterns = [
    r"\$\d{1,3}(,\d{3})*(\.\d{1,2})?",   # Format $11.1 ou $111,111.11
    r"\d+ dollars",                      # Format 11 dollars
    r"\d+ USD"                           # Format 11 USD
]
# <<<<<<<<<Money patterns for regex<<<<<<<<<
# >>>>>>>>>Destiny directory for news images>>>>>>>>>
directory = "output/"
excel_filename = "aljazeera_news.xlsx"
# <<<<<<<<<Destiny directory for news images<<<<<<<<<


class AljazeeraDatabaseService(Service):

    # >>>>>>>>>Store Aljazeera data in DataBase>>>>>>>>>
    def save_aljazeera_data(self, payload, data):
        try:
            for h in data:
                try:
                    # >>>>>>>>>Check if news item is already in DataBase>>>>>>>>>
                    item = AljazeeraModel.query.filter(and_(
                        AljazeeraModel.title==h[0],
                        AljazeeraModel.date==datetime.strptime(
                        h[3].replace("Last update ", ""),
                        "%d %b %Y"
                    ) if len(h) > 3 else None)).first()
                    # <<<<<<<<<Check if news item is already in DataBase<<<<<<<<<
                    # >>>>>>>>>Check if the image from the item have been downloaded correctly>>>>>>>>>
                    if item is not None:
                        if not item.picture_saved:
                            item.picture_saved = True if self.store_picture(directory=directory, id=item.id_aljazeera, url=h[4] if len(h) > 3 else h[2]) else False
                        continue
                    # <<<<<<<<<Check if the image from the item have been downloaded correctly<<<<<<<<<

                    # >>>>>>>>>Make filename of the excel file>>>>>>>>>
                    filename = AljazeeraModel.query.order_by(desc(AljazeeraModel.id_aljazeera)).first()
                    id_aljazeera = "1" if filename is None else filename.id_aljazeera + 1
                    filename = f"{id_aljazeera}-Aljazeera"
                    # <<<<<<<<<Make filename of the excel file<<<<<<<<<

                    # >>>>>>>>>Adding aljazeera data to Database>>>>>>>>>
                    aljazeera = AljazeeraModel()
                    aljazeera.id_aljazeera = int(id_aljazeera)
                    aljazeera.email = payload["email"]
                    aljazeera.search_phrase = payload["search_phrase"]
                    aljazeera.title = h[0]
                    aljazeera.date = datetime.strptime(
                        h[3].replace("Last update ", ""),
                        "%d %b %Y"
                    ) if len(h) > 3 else None
                    aljazeera.description = h[1]
                    aljazeera.picture_filename = filename
                    aljazeera.picture_url = h[4] if len(h) > 3 else h[2]
                    aljazeera.picture_saved = True if self.store_picture(directory=directory, id=str(id_aljazeera), url=h[4] if len(h) > 3 else h[2]) else False
                    aljazeera.count_search_phrase = h[0].upper().count(payload["search_phrase"].upper()) + h[1].upper().count(payload["search_phrase"].upper()) # Count of "search_phrase" in the title and description
                    aljazeera.money = True if any(re.search(pattern, h[0]) for pattern in money_patterns) or any(re.search(pattern, h[1]) for pattern in money_patterns) else False
                    aljazeera.dt_insert = datetime.now(pytz.timezone("America/Sao_Paulo"))
                    aljazeera.dt_update = datetime.now(pytz.timezone("America/Sao_Paulo"))
                    aljazeera.active = True
                    self.add(aljazeera)
                    # Save changes
                    self.commit_changes()
                    continue
                    # <<<<<<<<<Adding aljazeera data to Database<<<<<<<<<
                except IntegrityError as e:
                    db.session.rollback()
                    continue
            return {"status": True}
        except Exception as e:
            # >>>>>>>>>Tracing the Error>>>>>>>>>
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback_details = {
                "filename": exc_traceback.tb_frame.f_code.co_filename,
                "line_number": exc_traceback.tb_lineno,
                "function_name": exc_traceback.tb_frame.f_code.co_name,
                "exception_type": exc_type.__name__,
                "exception_message": str(exc_value)
            }
            print("=-==-==-=ERROR=-==-==-=")
            print(traceback_details)
            print("=-==-==-=ERROR=-==-==-=")
            # <<<<<<<<<Tracing the Error<<<<<<<<<
            return {"status": False, "msg": json.dumps(traceback_details)}
    # <<<<<<<<<Store Aljazeera data in DataBase<<<<<<<<<

    # >>>>>>>>>Save aljazeera data in an excel file>>>>>>>>>
    def aljazeera_excel_file(self):
        try:
            # Get items to fill excel file
            items = AljazeeraModel.query.filter(AljazeeraModel.active==True).all()
            data = [
                [
                    item.id_aljazeera, item.email, item.search_phrase, item.title,
                    item.date, item.description, item.picture_filename, item.picture_url,
                    item.count_search_phrase, item.money, item.dt_insert, item.dt_update
                ]
                for item in items
            ]

            # >>>>>>>>>Create a new workbook and select the first tab renaming it>>>>>>>>>
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Aljazeera-News"
            # <<<<<<<<<Create a new workbook and select the first tab renaming it<<<<<<<<<

            # >>>>>>>>>Set the header>>>>>>>>>
            header = [
                "ID_Aljazeera", "Email", "Search_Phrase", "Title",
                "Date", "Description", "Picture_Filename", "Picture_URL",
                "Count_Search_Phrase", "Money", "DT_Insert", "DT_Update"
            ]
            ws.append(header)
            # Freeze header
            ws.freeze_panes = "B2"
            # <<<<<<<<<Set the header<<<<<<<<<

            # >>>>>>>>>Format header (bold, black background, white font)>>>>>>>>>
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")  # White font color
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
            # <<<<<<<<<Format header (bold, black background, white font)<<<<<<<<<

            # >>>>>>>>>Insert data from the second line>>>>>>>>>
            for row in data:
                ws.append(row)
            # <<<<<<<<<Insert data from the second line<<<<<<<<<

            # >>>>>>>>>Center align and add borders to all cells>>>>>>>>>
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            # <<<<<<<<<Center align and add borders to all cells<<<<<<<<<

            # >>>>>>>>>Auto-adjust column widths based on content>>>>>>>>>
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obter a letra da coluna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Adicionar um pequeno buffer
                ws.column_dimensions[col_letter].width = adjusted_width
            # <<<<<<<<<Auto-adjust column widths based on content<<<<<<<<<

            # >>>>>>>>>Check if directory exists, otherwise create>>>>>>>>>
            os.makedirs(directory, exist_ok=True)
            # <<<<<<<<<Check if directory exists, otherwise create<<<<<<<<<

            # >>>>>>>>>Save the file to the specified path>>>>>>>>>
            file_path = os.path.join(directory, excel_filename)
            wb.save(file_path)
            # <<<<<<<<<Save the file to the specified path<<<<<<<<<
            return {"status": True}
        except Exception as e:
            # >>>>>>>>>Tracing the Error>>>>>>>>>
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback_details = {
                "filename": exc_traceback.tb_frame.f_code.co_filename,
                "line_number": exc_traceback.tb_lineno,
                "function_name": exc_traceback.tb_frame.f_code.co_name,
                "exception_type": exc_type.__name__,
                "exception_message": str(exc_value)
            }
            print("=-==-==-=ERROR=-==-==-=")
            print(traceback_details)
            print("=-==-==-=ERROR=-==-==-=")
            # <<<<<<<<<Tracing the Error<<<<<<<<<
            return {"status": False, "msg": json.dumps(traceback_details)}
    # <<<<<<<<<Save aljazeera data in an excel file<<<<<<<<<

    # >>>>>>>>>Send aljazeera excel by email>>>>>>>>>
    def aljazeera_excel_email(self, email):
        try:
            self.excel_by_email(os.path.join(directory, excel_filename), email, "Aljazeera Excel")
            return {"status": True}
        except Exception as e:
            # >>>>>>>>>Tracing the Error>>>>>>>>>
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback_details = {
                "filename": exc_traceback.tb_frame.f_code.co_filename,
                "line_number": exc_traceback.tb_lineno,
                "function_name": exc_traceback.tb_frame.f_code.co_name,
                "exception_type": exc_type.__name__,
                "exception_message": str(exc_value)
            }
            print("=-==-==-=ERROR=-==-==-=")
            print(traceback_details)
            print("=-==-==-=ERROR=-==-==-=")
            # <<<<<<<<<Tracing the Error<<<<<<<<<
            return {"status": False, "msg": json.dumps(traceback_details)}
    # <<<<<<<<<Send aljazeera excel by email<<<<<<<<<
